import io
import logging
import time
import xml.etree.ElementTree as ET
from datetime import datetime

import httpx
from openpyxl import load_workbook

from ..config import get_settings

logger = logging.getLogger(__name__)


def _auth() -> tuple[str, str]:
    s = get_settings()
    return (s.nextcloud_user, s.nextcloud_password)


def _webdav_url() -> str:
    s = get_settings()
    return s.nextcloud_url.rstrip("/") + s.nextcloud_file_path


# Short read-cache for preview flow; invalidated after every upload
_xlsx_cache: dict = {"data": None, "ts": 0.0, "etag": "", "last_modified": ""}
_XLSX_CACHE_TTL = 60  # seconds


async def download_xlsx(force: bool = False) -> bytes:
    if (
        not force
        and _xlsx_cache["data"] is not None
        and time.time() - _xlsx_cache["ts"] < _XLSX_CACHE_TTL
    ):
        logger.info("download_xlsx: returning cached xlsx (%d bytes, age=%.0fs)",
                    len(_xlsx_cache["data"]), time.time() - _xlsx_cache["ts"])
        return _xlsx_cache["data"]
    logger.info("download_xlsx: fetching from Nextcloud (force=%s)", force)
    async with httpx.AsyncClient(auth=_auth(), follow_redirects=True) as client:
        resp = await client.get(_webdav_url(), timeout=httpx.Timeout(10.0, read=60.0))
        resp.raise_for_status()
        data = resp.content
    logger.info("download_xlsx: downloaded %d bytes from Nextcloud etag=%s",
                len(data), resp.headers.get("etag", ""))
    _xlsx_cache["data"] = data
    _xlsx_cache["ts"] = time.time()
    _xlsx_cache["etag"] = resp.headers.get("etag", "")
    _xlsx_cache["last_modified"] = resp.headers.get("last-modified", "")
    return data


def _parse_propfind_meta(xml_text: str) -> dict:
    """Extract etag/last_modified/content_length from a PROPFIND multistatus XML body."""
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError as exc:
        logger.warning("_parse_propfind_meta: XML parse error: %s", exc)
        return {"etag": "", "last_modified": "", "content_length": 0}
    DAV = "DAV:"
    for prop in root.iter(f"{{{DAV}}}prop"):
        etag = (prop.findtext(f"{{{DAV}}}getetag") or "").strip()
        last_modified = (prop.findtext(f"{{{DAV}}}getlastmodified") or "").strip()
        cl_str = (prop.findtext(f"{{{DAV}}}getcontentlength") or "0").strip()
        try:
            content_length = int(cl_str)
        except ValueError:
            content_length = 0
        return {"etag": etag, "last_modified": last_modified, "content_length": content_length}
    return {"etag": "", "last_modified": "", "content_length": 0}


async def fetch_spreadsheet_meta() -> dict:
    """
    Return current server metadata (etag, last_modified, content_length) without
    downloading the file.  Tries HEAD first; falls back to PROPFIND if ETag is
    absent from the HEAD response (some WebDAV proxies strip it).
    """
    url = _webdav_url()
    auth = _auth()
    async with httpx.AsyncClient(auth=auth, follow_redirects=True) as client:
        # ── HEAD (preferred: no response body) ──────────────────────────────
        try:
            r = await client.head(url, timeout=httpx.Timeout(10.0))
            r.raise_for_status()
            if r.headers.get("etag"):
                logger.debug("fetch_spreadsheet_meta: HEAD ok etag=%s", r.headers.get("etag"))
                return {
                    "etag": r.headers.get("etag", ""),
                    "last_modified": r.headers.get("last-modified", ""),
                    "content_length": int(r.headers.get("content-length") or 0),
                }
            logger.debug("fetch_spreadsheet_meta: HEAD response missing etag — trying PROPFIND")
        except Exception as exc:
            logger.debug("fetch_spreadsheet_meta: HEAD failed (%s) — trying PROPFIND", exc)

        # ── PROPFIND fallback (standard WebDAV property query) ───────────────
        body = (
            '<?xml version="1.0" encoding="utf-8"?>'
            '<d:propfind xmlns:d="DAV:"><d:prop>'
            '<d:getetag/><d:getlastmodified/><d:getcontentlength/>'
            '</d:prop></d:propfind>'
        )
        r = await client.request(
            "PROPFIND", url,
            content=body.encode("utf-8"),
            headers={"Depth": "0", "Content-Type": "application/xml; charset=utf-8"},
            timeout=httpx.Timeout(10.0),
        )
        r.raise_for_status()
        meta = _parse_propfind_meta(r.text)
        logger.debug("fetch_spreadsheet_meta: PROPFIND ok etag=%s", meta.get("etag"))
        return meta


def get_cached_xlsx_meta() -> dict:
    """Return metadata captured from the last successful download_xlsx call."""
    return {
        "etag": _xlsx_cache.get("etag", ""),
        "last_modified": _xlsx_cache.get("last_modified", ""),
        "downloaded_at": (
            datetime.utcfromtimestamp(_xlsx_cache["ts"]).isoformat() + "Z"
            if _xlsx_cache["ts"] > 0 else None
        ),
        "size": len(_xlsx_cache["data"]) if _xlsx_cache["data"] else None,
    }


def _extract_row_color(ws, row_idx: int) -> str | None:
    """Return #RRGGBB from column A fill, or None if no significant color."""
    try:
        fill = ws.cell(row=row_idx, column=1).fill
        if fill and fill.fill_type == "solid":
            fg = fill.fgColor
            if fg and fg.type == "rgb":
                rgb = fg.rgb  # ARGB: 'FF4472C4'
                if len(rgb) == 8 and rgb[:2] == "FF":
                    hex6 = rgb[2:]
                    if hex6 not in ("000000", "FFFFFF", "ffffff"):
                        return "#" + hex6
    except Exception:
        pass
    return None


def _parse_sheet_rows(ws) -> list[dict]:
    """Parse one worksheet using the standard column mapping (B=ID, C=price, A=color)."""
    logger.info("_parse_sheet_rows: sheet='%s' max_row=%s max_col=%s",
                ws.title, ws.max_row, ws.max_column)
    items = []
    consecutive_empty = 0
    skipped_no_b = 0
    skipped_bad_id = 0

    for row_idx in range(3, 1001):
        col_a = ws.cell(row=row_idx, column=1).value
        col_b = ws.cell(row=row_idx, column=2).value
        col_c = ws.cell(row=row_idx, column=3).value

        if col_a is None and col_b is None and col_c is None:
            consecutive_empty += 1
            if consecutive_empty >= 30:
                logger.info("_parse_sheet_rows: sheet='%s' stopping at row %d (30 consecutive empty rows)",
                            ws.title, row_idx)
                break
            continue
        consecutive_empty = 0

        if col_b is None:
            skipped_no_b += 1
            continue
        try:
            pid = int(str(col_b).replace(",", "").strip())
        except (ValueError, TypeError):
            skipped_bad_id += 1
            logger.debug("_parse_sheet_rows: sheet='%s' row %d skipped — col B not an int: %r",
                         ws.title, row_idx, col_b)
            continue
        if pid <= 0:
            skipped_bad_id += 1
            continue

        if col_c is None or str(col_c).strip() == "":
            new_price = ""
        else:
            price_str = str(col_c).replace(",", "").strip()
            try:
                new_price = f"{float(price_str):.2f}"
            except (ValueError, TypeError):
                new_price = ""

        row_color = _extract_row_color(ws, row_idx)
        sheet_name = str(col_a).strip() if col_a is not None else ""
        items.append({"product_id": pid, "new_price": new_price, "row_color": row_color, "sheet_name": sheet_name})

    logger.info("_parse_sheet_rows: sheet='%s' parsed=%d skipped(no_b=%d bad_id=%d)",
                ws.title, len(items), skipped_no_b, skipped_bad_id)
    return items


def parse_price_list(xlsx_bytes: bytes) -> tuple[list[dict], list[dict]]:
    """
    Read ALL sheets from row 3 onward using the same column mapping.
    Column B = WooCommerce product ID, Column C = regular price, Column A = row color/name.
    If the same product ID appears in multiple sheets, the last sheet wins.
    Returns (items, duplicate_warnings).
    Each warning: {product_id, prev_sheet, final_sheet, prev_price, final_price}.
    """
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True)
    sheet_names = [ws.title for ws in wb.worksheets]
    logger.info("parse_price_list: xlsx=%d bytes, %d sheet(s): %s",
                len(xlsx_bytes), len(sheet_names), sheet_names)

    seen: dict[int, dict] = {}
    duplicates: list[dict] = []

    for ws in wb.worksheets:
        tab = ws.title
        sheet_items = _parse_sheet_rows(ws)
        logger.info("parse_price_list: sheet '%s' contributed %d product(s)", tab, len(sheet_items))
        for item in sheet_items:
            pid = item["product_id"]
            if pid in seen:
                duplicates.append({
                    "product_id": pid,
                    "prev_sheet": seen[pid].get("_tab", ""),
                    "final_sheet": tab,
                    "prev_price": seen[pid]["new_price"],
                    "final_price": item["new_price"],
                })
            item["_tab"] = tab
            seen[pid] = item

    wb.close()
    items = [{k: v for k, v in i.items() if k != "_tab"} for i in seen.values()]
    logger.info("parse_price_list: total unique products=%d duplicates=%d", len(items), len(duplicates))
    return items, duplicates


async def write_price_to_sheet(product_id: int, new_price: str) -> None:
    """Overwrite column C for the row whose column B matches product_id (searches all sheets)."""
    xlsx_bytes = await download_xlsx(force=True)
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes))
    logger.info("write_price_to_sheet: searching all %d sheet(s) for product_id=%d",
                len(wb.worksheets), product_id)

    found = False
    for ws in wb.worksheets:
        consecutive_empty = 0
        for row_idx in range(3, 1001):
            col_a = ws.cell(row=row_idx, column=1).value
            col_b = ws.cell(row=row_idx, column=2).value
            col_c = ws.cell(row=row_idx, column=3).value
            if col_a is None and col_b is None and col_c is None:
                consecutive_empty += 1
                if consecutive_empty >= 30:
                    break
                continue
            consecutive_empty = 0
            if col_b is None:
                continue
            try:
                pid = int(str(col_b).replace(",", "").strip())
            except (ValueError, TypeError):
                continue
            if pid == product_id:
                try:
                    ws.cell(row=row_idx, column=3).value = float(new_price) if new_price else None
                except (ValueError, TypeError):
                    ws.cell(row=row_idx, column=3).value = new_price or None
                logger.info("write_price_to_sheet: updated sheet='%s' row=%d product_id=%d new_price=%s",
                            ws.title, row_idx, product_id, new_price)
                found = True
                break
        if found:
            break

    if not found:
        logger.warning("write_price_to_sheet: product_id=%d not found in any sheet", product_id)

    await _upload_wb(wb)


async def write_back_to_sheet(results: list[dict]) -> None:
    """Update columns E (status), F (sync time), G (error) by product_id (column B)."""
    result_map = {r["product_id"]: r for r in results}

    xlsx_bytes = await download_xlsx(force=True)
    wb = load_workbook(filename=io.BytesIO(xlsx_bytes))
    ws = wb.active

    consecutive_empty = 0
    for row_idx in range(3, 1001):
        col_a = ws.cell(row=row_idx, column=1).value
        col_b = ws.cell(row=row_idx, column=2).value
        col_c = ws.cell(row=row_idx, column=3).value
        if col_a is None and col_b is None and col_c is None:
            consecutive_empty += 1
            if consecutive_empty >= 30:
                break
            continue
        consecutive_empty = 0
        if col_b is None:
            continue
        try:
            pid = int(str(col_b).replace(",", "").strip())
        except (ValueError, TypeError):
            continue
        if pid not in result_map:
            continue
        r = result_map[pid]
        ws.cell(row=row_idx, column=5).value = r.get("status", "")
        ws.cell(row=row_idx, column=6).value = r.get("synced_at", "")
        ws.cell(row=row_idx, column=7).value = r.get("error_message", "")

    await _upload_wb(wb)


async def _upload_wb(wb) -> None:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    async with httpx.AsyncClient(auth=_auth(), follow_redirects=True) as client:
        resp = await client.put(
            _webdav_url(),
            content=buf.read(),
            timeout=60,
            headers={
                "Content-Type": (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            },
        )
        resp.raise_for_status()
    _xlsx_cache["data"] = None  # invalidate so next read fetches the just-uploaded file
