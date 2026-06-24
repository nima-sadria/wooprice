"""
NextcloudSourceAdapter — A2.2 initial adapter implementation.

Reads a WooPrice-format XLSX file from a Nextcloud / OnlyOffice WebDAV source,
validates structural integrity, generates a SourceSnapshot, produces per-row
SourceRowProvenance, and streams SourceRow objects.

Adapter responsibility ends at: Source → Validation → Snapshot → Provenance → Row Streaming.
"""
from __future__ import annotations

import hashlib
import io
import json
import logging
import xml.etree.ElementTree as ET
from collections.abc import AsyncIterator
from datetime import datetime, timezone
from typing import Optional
from uuid import uuid4

import httpx
from openpyxl import load_workbook

from ..base import SourceAdapter, SourceRow, SourceValidationResult, hash_row
from ..capabilities import SourceCapabilities
from ..checkpoint import SourceCheckpoint
from ..provenance import SourceRowProvenance
from ..snapshot import SourceSnapshot

logger = logging.getLogger(__name__)

# Column layout for the WooPrice spreadsheet format.
# Row 1: header row (used for schema_hash computation across all sheets).
# Row 2: reserved (formatting / sub-header). Data starts at row 3.
# Column A: descriptive label / sheet name echo
# Column B: product identifier (stable integer key)
# Column C: price (raw, may be empty or non-numeric)
_DATA_START_ROW = 3
_COL_LABEL = 1   # A
_COL_ID = 2      # B — stable row identity
_COL_PRICE = 3   # C


class NextcloudSourceAdapter(SourceAdapter):
    """
    Reads a WooPrice XLSX spreadsheet from Nextcloud / OnlyOffice via WebDAV.

    Lifecycle:
        1. connect()           — download file + capture metadata
        2. validate_source()   — structural validation (no duplicates, valid IDs)
        3. fetch_snapshot()    — create immutable snapshot descriptor; binds streaming context
        4. stream_rows()       — yield rows with provenance tied to the generated snapshot
        5. get_checkpoint()    — return ETag-based checkpoint
        6. advance_checkpoint(cp, db) — persist checkpoint durably (requires db session)

    Configuration is injected at construction time so the adapter is independently
    testable without global settings.
    """

    SOURCE_TYPE = "nextcloud_xlsx"

    def __init__(
        self,
        *,
        source_id: str,
        url: str,
        username: str,
        password: str,
        file_path: str,
    ) -> None:
        self._source_id = source_id
        self._url = url.rstrip("/")
        self._username = username
        self._password = password
        self._file_path = file_path
        self._xlsx_bytes: Optional[bytes] = None
        self._meta: dict = {}
        self._current_snapshot: Optional[SourceSnapshot] = None

    # ── Internal helpers ──────────────────────────────────────────────────────

    def _webdav_url(self) -> str:
        return self._url + self._file_path

    def _auth(self) -> tuple[str, str]:
        return (self._username, self._password)

    async def _fetch_file_and_meta(self) -> tuple[bytes, dict]:
        """Download the XLSX file and capture WebDAV metadata in a single GET."""
        async with httpx.AsyncClient(
            auth=self._auth(), follow_redirects=True
        ) as client:
            resp = await client.get(
                self._webdav_url(),
                timeout=httpx.Timeout(10.0, read=60.0),
            )
            resp.raise_for_status()
        meta = {
            "etag": resp.headers.get("etag", ""),
            "last_modified": resp.headers.get("last-modified", ""),
        }
        return resp.content, meta

    async def _fetch_meta_only(self) -> dict:
        """Retrieve server metadata without downloading the file body."""
        url = self._webdav_url()
        async with httpx.AsyncClient(
            auth=self._auth(), follow_redirects=True
        ) as client:
            try:
                r = await client.head(url, timeout=httpx.Timeout(10.0))
                r.raise_for_status()
                if r.headers.get("etag"):
                    return {
                        "etag": r.headers.get("etag", ""),
                        "last_modified": r.headers.get("last-modified", ""),
                    }
            except Exception:
                pass

            # PROPFIND fallback
            body = (
                '<?xml version="1.0" encoding="utf-8"?>'
                '<d:propfind xmlns:d="DAV:"><d:prop>'
                "<d:getetag/><d:getlastmodified/>"
                "</d:prop></d:propfind>"
            )
            r = await client.request(
                "PROPFIND",
                url,
                content=body.encode("utf-8"),
                headers={"Depth": "0", "Content-Type": "application/xml; charset=utf-8"},
                timeout=httpx.Timeout(10.0),
            )
            r.raise_for_status()
            return _parse_propfind_meta(r.text)

    # ── Static row parsing ────────────────────────────────────────────────────

    @staticmethod
    def _parse_xlsx_rows(xlsx_bytes: bytes) -> tuple[list[dict], list[str], str]:
        """
        Parse all sheets and return (rows, errors, schema_hash).

        rows: list of dicts with keys product_id, price_raw, label, sheet_name
        errors: list of validation error messages (duplicates, missing IDs)
        schema_hash: SHA-256 over header rows of ALL worksheets

        All rows in all sheets are inspected — there is no silent row cap.
        Consecutive-empty-row early-exit is applied per-sheet only to skip
        trailing whitespace, but does not hide any rows with data following gaps.
        """
        wb = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True)
        errors: list[str] = []
        seen_ids: dict[str, str] = {}  # product_id_str -> sheet_name
        all_rows: list[dict] = []

        # Build schema_hash from header row of every worksheet.
        all_headers: list[list[str]] = []
        for ws in wb.worksheets:
            headers = [
                str(ws.cell(row=1, column=c).value or "")
                for c in range(1, ws.max_column + 1)
            ]
            all_headers.append(headers)

        schema_hash = hashlib.sha256(
            json.dumps(all_headers).encode()
        ).hexdigest() if all_headers else hashlib.sha256(b"empty").hexdigest()

        for ws in wb.worksheets:
            sheet_name = ws.title

            for row_idx in range(_DATA_START_ROW, ws.max_row + 1):
                col_a = ws.cell(row=row_idx, column=_COL_LABEL).value
                col_b = ws.cell(row=row_idx, column=_COL_ID).value
                col_c = ws.cell(row=row_idx, column=_COL_PRICE).value

                if col_a is None and col_b is None and col_c is None:
                    continue

                if col_b is None:
                    errors.append(
                        f"Row {row_idx} in sheet '{sheet_name}' "
                        "has no product identifier (column B is empty)."
                    )
                    continue

                try:
                    pid_int = int(str(col_b).replace(",", "").strip())
                except (ValueError, TypeError):
                    errors.append(
                        f"Row {row_idx} in sheet '{sheet_name}' "
                        f"has a non-integer product identifier: {col_b!r}."
                    )
                    continue

                if pid_int <= 0:
                    errors.append(
                        f"Row {row_idx} in sheet '{sheet_name}' "
                        f"has an invalid product identifier {pid_int} (must be > 0)."
                    )
                    continue

                row_ref = str(pid_int)
                if row_ref in seen_ids:
                    errors.append(
                        f"Duplicate product identifier {row_ref} found in "
                        f"sheet '{sheet_name}' (first seen in '{seen_ids[row_ref]}')."
                    )
                    continue

                seen_ids[row_ref] = sheet_name
                all_rows.append(
                    {
                        "product_id": pid_int,
                        "price_raw": "" if col_c is None else str(col_c).strip(),
                        "label": "" if col_a is None else str(col_a).strip(),
                        "sheet_name": sheet_name,
                    }
                )

        wb.close()
        return all_rows, errors, schema_hash

    # ── SourceAdapter interface ───────────────────────────────────────────────

    async def connect(self) -> None:
        """Download the source file and capture metadata."""
        self._xlsx_bytes, self._meta = await self._fetch_file_and_meta()
        logger.info(
            "NextcloudSourceAdapter.connect: source_id=%s bytes=%d etag=%s",
            self._source_id,
            len(self._xlsx_bytes),
            self._meta.get("etag", ""),
        )

    async def validate_source(self) -> SourceValidationResult:
        """
        Validate structure: reachable, readable, stable identifiers present, no duplicates.

        Duplicate product identifiers are validation ERRORS, not warnings.
        connect() must be called before validate_source().
        """
        if self._xlsx_bytes is None:
            return SourceValidationResult(
                is_valid=False,
                errors=["connect() must be called before validate_source()."],
            )

        _, errors, _ = self._parse_xlsx_rows(self._xlsx_bytes)
        return SourceValidationResult(is_valid=len(errors) == 0, errors=errors)

    async def fetch_snapshot(self) -> SourceSnapshot:
        """
        Build an immutable snapshot descriptor and bind it as the streaming context.

        connect() must be called before fetch_snapshot().
        stream_rows() uses the most recently generated snapshot; always call
        fetch_snapshot() before stream_rows().
        """
        if self._xlsx_bytes is None:
            raise RuntimeError("connect() must be called before fetch_snapshot().")

        rows, errors, schema_hash = self._parse_xlsx_rows(self._xlsx_bytes)
        if errors:
            raise ValueError(
                f"Source has {len(errors)} validation error(s); "
                "snapshot cannot be created for an invalid source."
            )

        fingerprint = hashlib.sha256(
            json.dumps(
                {
                    "etag": self._meta.get("etag", ""),
                    "last_modified": self._meta.get("last_modified", ""),
                    "content_sha256": hashlib.sha256(self._xlsx_bytes).hexdigest(),
                },
                sort_keys=True,
            ).encode()
        ).hexdigest()

        snapshot = SourceSnapshot(
            snapshot_id=str(uuid4()),
            source_id=self._source_id,
            created_at=datetime.now(tz=timezone.utc),
            schema_hash=schema_hash,
            row_count=len(rows),
            source_fingerprint=fingerprint,
        )
        self._current_snapshot = snapshot
        return snapshot

    async def stream_rows(self) -> AsyncIterator[SourceRow]:  # type: ignore[override]
        """
        Async generator yielding SourceRow objects with provenance bound to the
        adapter-generated snapshot from the most recent fetch_snapshot() call.

        fetch_snapshot() must be called before stream_rows(). This ensures row
        provenance references an adapter-generated snapshot, not an arbitrary
        external ID.
        """
        if self._xlsx_bytes is None:
            raise RuntimeError("connect() must be called before stream_rows().")
        if self._current_snapshot is None:
            raise RuntimeError("fetch_snapshot() must be called before stream_rows().")

        rows, errors, _ = self._parse_xlsx_rows(self._xlsx_bytes)
        if errors:
            raise ValueError(
                f"Cannot stream rows from an invalid source ({len(errors)} error(s))."
            )

        snapshot_id = self._current_snapshot.snapshot_id
        for row_data in rows:
            row_ref = str(row_data["product_id"])
            rh = hash_row(row_data)
            provenance = SourceRowProvenance(
                source_id=self._source_id,
                source_row_ref=row_ref,
                source_snapshot_id=snapshot_id,
                source_row_hash=rh,
            )
            yield SourceRow(
                row_ref=row_ref,
                raw_data=row_data,
                row_hash=rh,
                provenance=provenance,
            )

    def get_capabilities(self) -> SourceCapabilities:
        return SourceCapabilities(
            supports_streaming=False,   # XLSX adapter reads full workbook into memory
            supports_checkpointing=True,
            supports_deletions=False,
            supports_incremental_sync=False,
            supports_snapshots=True,
        )

    async def get_checkpoint(self) -> Optional[SourceCheckpoint]:
        """
        Return a checkpoint based on the WebDAV ETag.

        Requires connect() to have been called so the ETag is captured.
        Falls back to a PROPFIND if the ETag was not available at connect time.
        """
        etag = self._meta.get("etag", "")
        if not etag:
            try:
                meta = await self._fetch_meta_only()
                etag = meta.get("etag", "")
            except Exception:
                return None

        if not etag:
            return None

        return SourceCheckpoint(
            source_id=self._source_id,
            checkpoint_value=etag,
            checkpointed_at=datetime.now(tz=timezone.utc),
            checkpoint_type="etag",
        )

    async def advance_checkpoint(
        self,
        checkpoint: SourceCheckpoint,
        db=None,
    ) -> None:
        """
        Persist checkpoint durably after a successful sync cycle.

        If db (a SQLAlchemy Session) is provided the checkpoint is persisted
        transactionally via CheckpointRepository. If db is None the checkpoint
        is logged only (useful in contexts where no DB session is available).
        """
        if db is not None:
            from ...repositories.checkpoint_repository import CheckpointRepository
            repo = CheckpointRepository(db)
            repo.save(checkpoint)
            db.commit()
            logger.info(
                "NextcloudSourceAdapter.advance_checkpoint: persisted source_id=%s value=%s",
                self._source_id,
                checkpoint.checkpoint_value,
            )
        else:
            logger.warning(
                "NextcloudSourceAdapter.advance_checkpoint: no db session provided; "
                "checkpoint NOT persisted. source_id=%s value=%s",
                self._source_id,
                checkpoint.checkpoint_value,
            )


# ── Utilities ─────────────────────────────────────────────────────────────────

def _parse_propfind_meta(xml_text: str) -> dict:
    try:
        root = ET.fromstring(xml_text)
    except ET.ParseError:
        return {"etag": "", "last_modified": ""}
    DAV = "DAV:"
    for prop in root.iter(f"{{{DAV}}}prop"):
        return {
            "etag": (prop.findtext(f"{{{DAV}}}getetag") or "").strip(),
            "last_modified": (prop.findtext(f"{{{DAV}}}getlastmodified") or "").strip(),
        }
    return {"etag": "", "last_modified": ""}
