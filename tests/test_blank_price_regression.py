"""Regression tests for the blank-price-classified-as-invalid production bug.

Business rule: a blank or zero sheet price means "out of stock intent" and must
never be treated as a critical/invalid price. Only genuine non-numeric garbage
(e.g. "abc", "12xx", "--100") is a true parse failure and must remain invalid
and block apply.

Run directly (no pytest dependency):
    python tests/test_blank_price_regression.py
"""
import os
import sys
from types import SimpleNamespace

os.environ.setdefault("NEXTCLOUD_URL", "http://example.invalid")
os.environ.setdefault("NEXTCLOUD_USER", "x")
os.environ.setdefault("NEXTCLOUD_PASSWORD", "x")
os.environ.setdefault("NEXTCLOUD_FILE_PATH", "/x.xlsx")
os.environ.setdefault("WC_URL", "http://example.invalid")
os.environ.setdefault("WC_KEY", "x")
os.environ.setdefault("WC_SECRET", "x")
os.environ.setdefault("DATABASE_URL", "sqlite:///:memory:")

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook  # noqa: E402

from app.main import _classify_row, _compute_dry_run_summary  # noqa: E402
from app.services.nextcloud import _parse_sheet_rows  # noqa: E402


def make_item(**kw):
    defaults = dict(
        product_id=1, product_name="Test", new_price="", old_price=None,
        stock_status="instock", change_status=None,
        price_changed=0, stock_changed=0, missing_image=0, missing_cost=0,
    )
    defaults.update(kw)
    return SimpleNamespace(**defaults)


def build_ws(rows):
    """rows: list of (col_a, col_b, col_c) tuples, written starting at row 3."""
    wb = Workbook()
    ws = wb.active
    for i, (a, b, c) in enumerate(rows):
        r = 3 + i
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        ws.cell(row=r, column=3, value=c)
    return ws


def test_parser_blank_vs_garbage():
    ws = build_ws([
        ("Cat", 101, None),     # blank -> out-of-stock intent
        ("Cat", 102, "  "),     # whitespace-only -> blank
        ("Cat", 103, "abc"),    # garbage -> true parse error
        ("Cat", 104, "12xx"),   # garbage
        ("Cat", 105, "--100"),  # garbage
        ("Cat", 106, 19.5),     # valid numeric
        ("Cat", 107, "0"),      # "0" is a recognized out-of-stock marker -> normalized to blank
    ])
    items = {i["product_id"]: i for i in _parse_sheet_rows(ws)}

    assert items[101]["new_price"] == "" and items[101]["price_parse_error"] is False
    assert items[102]["new_price"] == "" and items[102]["price_parse_error"] is False
    assert items[103]["price_parse_error"] is True and items[103]["new_price"] == "abc"
    assert items[104]["price_parse_error"] is True and items[104]["new_price"] == "12xx"
    assert items[105]["price_parse_error"] is True and items[105]["new_price"] == "--100"
    assert items[106]["price_parse_error"] is False and items[106]["new_price"] == "19.50"
    assert items[107]["price_parse_error"] is False and items[107]["new_price"] == ""
    print("test_parser_blank_vs_garbage: PASS")


def test_classify_row_blank_price_not_invalid():
    wc = {"price": "50.00", "stock_status": "instock"}
    clf = _classify_row(1, "", wc, last_price_updated=None, cache_row=None, price_parse_error=False)
    assert clf["change_status"] != "invalid", clf
    assert clf["change_status"] in ("new", "changed")
    assert clf["stock_changed"] == 1
    print("test_classify_row_blank_price_not_invalid: PASS")


def test_classify_row_blank_price_unchanged_when_already_outofstock():
    wc = {"price": "", "stock_status": "outofstock"}
    clf = _classify_row(2, "", wc, last_price_updated="2024-01-01", cache_row=None, price_parse_error=False)
    assert clf["change_status"] == "unchanged", clf
    print("test_classify_row_blank_price_unchanged_when_already_outofstock: PASS")


def test_classify_row_zero_price_same_as_blank():
    wc = {"price": "50.00", "stock_status": "instock"}
    clf = _classify_row(3, "0.00", wc, last_price_updated="2024-01-01", cache_row=None, price_parse_error=False)
    assert clf["change_status"] == "changed", clf
    assert clf["stock_changed"] == 1
    print("test_classify_row_zero_price_same_as_blank: PASS")


def test_classify_row_garbage_still_invalid():
    wc = {"price": "50.00", "stock_status": "instock"}
    clf = _classify_row(4, "abc", wc, last_price_updated=None, cache_row=None, price_parse_error=True)
    assert clf["change_status"] == "invalid", clf
    print("test_classify_row_garbage_still_invalid: PASS")


def test_dry_run_summary_blank_price_is_warning_not_critical():
    item_blank = make_item(product_id=10, new_price="", old_price="50.00",
                            change_status="changed", stock_changed=1, price_changed=1,
                            stock_status="instock")
    summary = _compute_dry_run_summary([item_blank], alarm_threshold=float("inf"))
    assert summary["critical_errors"] == [], summary["critical_errors"]
    assert any(w["type"] == "out_of_stock_marker" for w in summary["warnings"]), summary["warnings"]
    assert summary["stock_to_outofstock"] == 1
    assert summary["dry_run_status"] == "warnings"
    print("test_dry_run_summary_blank_price_is_warning_not_critical: PASS")


def test_dry_run_summary_garbage_price_is_critical_and_blocks():
    item_garbage = make_item(product_id=11, new_price="abc", old_price="50.00",
                              change_status="invalid")
    summary = _compute_dry_run_summary([item_garbage], alarm_threshold=float("inf"))
    assert any(e["type"] == "invalid_price" for e in summary["critical_errors"]), summary["critical_errors"]
    assert summary["dry_run_status"] == "blocked"
    print("test_dry_run_summary_garbage_price_is_critical_and_blocks: PASS")


def test_dry_run_summary_mixed_does_not_block_on_blank_alone():
    item_blank = make_item(product_id=20, new_price="", old_price="50.00",
                            change_status="changed", stock_changed=1, price_changed=1)
    summary = _compute_dry_run_summary([item_blank], alarm_threshold=float("inf"))
    assert summary["dry_run_status"] != "blocked", summary
    print("test_dry_run_summary_mixed_does_not_block_on_blank_alone: PASS")


def test_invalid_count_excludes_blank_rows():
    items_clf = [
        _classify_row(1, "", {"price": "10", "stock_status": "instock"}, None, None, price_parse_error=False),
        _classify_row(2, "abc", {"price": "10", "stock_status": "instock"}, None, None, price_parse_error=True),
    ]
    invalid_count = sum(1 for c in items_clf if c["change_status"] == "invalid")
    assert invalid_count == 1, invalid_count
    print("test_invalid_count_excludes_blank_rows: PASS")


def test_parser_out_of_stock_markers_not_invalid():
    ws = build_ws([
        ("Cat", 201, "-"),
        ("Cat", 202, "ناموجود"),
        ("Cat", 203, "ناموجود شد"),
        ("Cat", 204, "تماس بگیرید"),
        ("Cat", 205, "out of stock"),
        ("Cat", 206, "OOS"),
        ("Cat", 207, "N/A"),
        ("Cat", 208, "na"),
        ("Cat", 209, "0.00"),
        ("Cat", 210, "x"),
        ("Cat", 211, "❌"),
        ("Cat", 212, "✗"),
        ("Cat", 213, "×"),
    ])
    items = {i["product_id"]: i for i in _parse_sheet_rows(ws)}
    for pid in (201, 202, 203, 204, 205, 206, 207, 208, 209, 210, 211, 212, 213):
        assert items[pid]["new_price"] == "", (pid, items[pid])
        assert items[pid]["price_parse_error"] is False, (pid, items[pid])
    print("test_parser_out_of_stock_markers_not_invalid: PASS")


def test_parser_persian_and_arabic_numerics():
    ws = build_ws([
        ("Cat", 301, "۱۲۳۴۵۶"),     # Persian digits
        ("Cat", 302, "۱۲۳,۴۵۶"),    # Persian digits + comma separator
        ("Cat", 303, "123,456"),    # ASCII comma separator
        ("Cat", 304, "123٬456"),    # Arabic thousands separator (U+066C)
    ])
    items = {i["product_id"]: i for i in _parse_sheet_rows(ws)}
    for pid in (301, 302, 303, 304):
        assert items[pid]["price_parse_error"] is False, (pid, items[pid])
        assert items[pid]["new_price"] == "123456.00", (pid, items[pid])
    print("test_parser_persian_and_arabic_numerics: PASS")


def test_parser_true_garbage_still_invalid():
    ws = build_ws([
        ("Cat", 401, "abc"),
        ("Cat", 402, "12xx"),
        ("Cat", 403, "--100"),
        ("Cat", 404, "random text"),
    ])
    items = {i["product_id"]: i for i in _parse_sheet_rows(ws)}
    for pid in (401, 402, 403, 404):
        assert items[pid]["price_parse_error"] is True, (pid, items[pid])
    print("test_parser_true_garbage_still_invalid: PASS")


def test_classify_row_out_of_stock_marker_strings_not_invalid():
    wc = {"price": "50.00", "stock_status": "instock"}
    for marker_text in ("-", "ناموجود", "تماس بگیرید", "out of stock", "n/a", "x", "❌", "✗", "×"):
        ws = build_ws([("Cat", 1, marker_text)])
        parsed = _parse_sheet_rows(ws)[0]
        clf = _classify_row(1, parsed["new_price"], wc, last_price_updated="2024-01-01",
                             cache_row=None, price_parse_error=parsed["price_parse_error"])
        assert clf["change_status"] != "invalid", (marker_text, clf)
        assert clf["change_status"] == "changed", (marker_text, clf)
        assert clf["stock_changed"] == 1, (marker_text, clf)
    print("test_classify_row_out_of_stock_marker_strings_not_invalid: PASS")


if __name__ == "__main__":
    test_parser_blank_vs_garbage()
    test_classify_row_blank_price_not_invalid()
    test_classify_row_blank_price_unchanged_when_already_outofstock()
    test_classify_row_zero_price_same_as_blank()
    test_classify_row_garbage_still_invalid()
    test_dry_run_summary_blank_price_is_warning_not_critical()
    test_dry_run_summary_garbage_price_is_critical_and_blocks()
    test_dry_run_summary_mixed_does_not_block_on_blank_alone()
    test_invalid_count_excludes_blank_rows()
    test_parser_out_of_stock_markers_not_invalid()
    test_parser_persian_and_arabic_numerics()
    test_parser_true_garbage_still_invalid()
    test_classify_row_out_of_stock_marker_strings_not_invalid()
    print("ALL TESTS PASSED")
