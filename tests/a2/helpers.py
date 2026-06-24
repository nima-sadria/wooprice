"""Shared test helpers for A2 tests."""
import io

from openpyxl import Workbook


def _make_xlsx(rows: list[tuple]) -> bytes:
    """
    Build a minimal WooPrice-format XLSX in memory.

    rows: list of (col_a, col_b, col_c) tuples, inserted starting at row 3.
    Rows 1–2 are left as header placeholders.
    """
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "Label"
    ws.cell(row=1, column=2).value = "Product ID"
    ws.cell(row=1, column=3).value = "Price"
    for i, (a, b, c) in enumerate(rows, start=3):
        ws.cell(row=i, column=1).value = a
        ws.cell(row=i, column=2).value = b
        ws.cell(row=i, column=3).value = c
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_xlsx_multisheet(sheets: list[list[tuple]]) -> bytes:
    """
    Build a WooPrice-format XLSX with multiple worksheets.

    sheets: list of row-lists, one per sheet.
    Each sheet row is (col_a, col_b, col_c), inserted at data rows starting from row 3.
    The first sheet uses the default active sheet; additional sheets are appended.
    """
    wb = Workbook()
    for i, rows in enumerate(sheets):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.cell(row=1, column=1).value = "Label"
        ws.cell(row=1, column=2).value = "Product ID"
        ws.cell(row=1, column=3).value = "Price"
        for j, (a, b, c) in enumerate(rows, start=3):
            ws.cell(row=j, column=1).value = a
            ws.cell(row=j, column=2).value = b
            ws.cell(row=j, column=3).value = c
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_xlsx_with_late_duplicate(early_pid: int, late_row: int) -> bytes:
    """
    Build a single-sheet XLSX where product `early_pid` appears at both
    data row 3 and the specified `late_row` (1-based physical row index).

    Used to verify that duplicate detection is not silently capped at a row limit.
    """
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "Label"
    ws.cell(row=1, column=2).value = "Product ID"
    ws.cell(row=1, column=3).value = "Price"
    # Early occurrence
    ws.cell(row=3, column=1).value = "First"
    ws.cell(row=3, column=2).value = early_pid
    ws.cell(row=3, column=3).value = "100"
    # Late duplicate at the specified high row number
    ws.cell(row=late_row, column=1).value = "Duplicate"
    ws.cell(row=late_row, column=2).value = early_pid
    ws.cell(row=late_row, column=3).value = "200"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
