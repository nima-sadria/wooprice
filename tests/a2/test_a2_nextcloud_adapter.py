"""
Integration tests — NextcloudSourceAdapter.

All HTTP calls are mocked; no live Nextcloud instance required.
"""
import os
os.environ.setdefault("A2_DATABASE_URL", "sqlite:///:memory:")
os.environ.setdefault("DATABASE_URL", "sqlite:///:memory:")
os.environ.setdefault("NEXTCLOUD_URL", "http://example.invalid")
os.environ.setdefault("NEXTCLOUD_USER", "x")
os.environ.setdefault("NEXTCLOUD_PASSWORD", "x")
os.environ.setdefault("NEXTCLOUD_FILE_PATH", "/x.xlsx")
os.environ.setdefault("WC_URL", "http://example.invalid")
os.environ.setdefault("WC_KEY", "x")
os.environ.setdefault("WC_SECRET", "x")

import asyncio
from datetime import datetime, timezone
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from app.a2.sources.adapters.nextcloud import NextcloudSourceAdapter
from app.a2.sources.checkpoint import SourceCheckpoint
from tests.a2.helpers import _make_xlsx, _make_xlsx_multisheet, _make_xlsx_with_late_duplicate


def _adapter(**kwargs):
    defaults = dict(
        source_id="test-src",
        url="http://nc.example.invalid",
        username="user",
        password="pass",
        file_path="/prices.xlsx",
    )
    defaults.update(kwargs)
    return NextcloudSourceAdapter(**defaults)


def _mock_response(xlsx_bytes: bytes, etag: str = '"etag-abc"'):
    resp = MagicMock()
    resp.content = xlsx_bytes
    resp.headers = {"etag": etag, "last-modified": "Tue, 24 Jun 2026 00:00:00 GMT"}
    resp.raise_for_status = MagicMock()
    return resp


def _run(coro):
    return asyncio.run(coro)


async def _collect_stream(adapter):
    rows = []
    async for row in adapter.stream_rows():
        rows.append(row)
    return rows


# ── connect() ─────────────────────────────────────────────────────────────────

def test_connect_downloads_file():
    xlsx = _make_xlsx([("Prod A", 101, "50000"), ("Prod B", 102, "80000")])
    adapter = _adapter()
    with patch.object(adapter, "_fetch_file_and_meta", new=AsyncMock(return_value=(xlsx, {"etag": '"e1"'}))):
        _run(adapter.connect())
    assert adapter._xlsx_bytes == xlsx
    assert adapter._meta["etag"] == '"e1"'


# ── validate_source() ─────────────────────────────────────────────────────────

def test_validate_valid_source():
    xlsx = _make_xlsx([("A", 1, "100"), ("B", 2, "200")])
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {"etag": '"e"'}
    result = _run(adapter.validate_source())
    assert result.is_valid is True
    assert result.errors == []


def test_validate_fails_before_connect():
    adapter = _adapter()
    result = _run(adapter.validate_source())
    assert result.is_valid is False
    assert any("connect()" in e for e in result.errors)


def test_validate_detects_duplicate_ids():
    xlsx = _make_xlsx([("A", 10, "100"), ("B", 10, "200")])
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {}
    result = _run(adapter.validate_source())
    assert result.is_valid is False
    assert any("10" in e for e in result.errors)


def test_validate_duplicate_is_error_not_warning():
    xlsx = _make_xlsx([("A", 99, "100"), ("B", 99, "200")])
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {}
    result = _run(adapter.validate_source())
    assert result.is_valid is False
    assert result.warnings == []


def test_validate_missing_product_id_is_error():
    xlsx = _make_xlsx([("A", None, "100"), ("B", 2, "200")])
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {}
    result = _run(adapter.validate_source())
    assert result.is_valid is False
    assert any("no product identifier" in e.lower() for e in result.errors)


def test_validate_non_integer_id_is_error():
    xlsx = _make_xlsx([("A", "not-an-id", "100")])
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {}
    result = _run(adapter.validate_source())
    assert result.is_valid is False


# ── fetch_snapshot() ──────────────────────────────────────────────────────────

def test_fetch_snapshot_valid_source():
    xlsx = _make_xlsx([("A", 1, "100"), ("B", 2, "200")])
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {"etag": '"e"', "last_modified": ""}
    snap = _run(adapter.fetch_snapshot())
    assert snap.source_id == "test-src"
    assert snap.row_count == 2
    assert len(snap.snapshot_id) == 36  # UUID
    assert snap.schema_hash
    assert snap.source_fingerprint


def test_fetch_snapshot_raises_on_invalid_source():
    xlsx = _make_xlsx([("A", 5, "100"), ("B", 5, "200")])  # duplicate
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {}
    with pytest.raises(ValueError, match="validation error"):
        _run(adapter.fetch_snapshot())


def test_fetch_snapshot_raises_before_connect():
    adapter = _adapter()
    with pytest.raises(RuntimeError, match="connect()"):
        _run(adapter.fetch_snapshot())


def test_snapshot_ids_are_unique():
    xlsx = _make_xlsx([("A", 1, "100")])
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {"etag": '"e"', "last_modified": ""}
    s1 = _run(adapter.fetch_snapshot())
    s2 = _run(adapter.fetch_snapshot())
    assert s1.snapshot_id != s2.snapshot_id


def test_fetch_snapshot_stores_snapshot_internally():
    """fetch_snapshot() must bind the snapshot so stream_rows() can use it."""
    xlsx = _make_xlsx([("A", 1, "100")])
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {"etag": '"e"', "last_modified": ""}
    assert adapter._current_snapshot is None
    snap = _run(adapter.fetch_snapshot())
    assert adapter._current_snapshot is snap


# ── stream_rows() — public contract ───────────────────────────────────────────

def test_stream_rows_yields_correct_count():
    xlsx = _make_xlsx([("A", 10, "100"), ("B", 20, "200"), ("C", 30, "300")])
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {"etag": '"e"', "last_modified": ""}
    _run(adapter.fetch_snapshot())
    rows = asyncio.run(_collect_stream(adapter))
    assert len(rows) == 3


def test_stream_rows_stable_row_ref():
    xlsx = _make_xlsx([("A", 42, "100")])
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {"etag": '"e"', "last_modified": ""}
    _run(adapter.fetch_snapshot())
    rows = asyncio.run(_collect_stream(adapter))
    assert rows[0].row_ref == "42"


def test_stream_rows_provenance_attached():
    xlsx = _make_xlsx([("A", 42, "100")])
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {"etag": '"e"', "last_modified": ""}
    snapshot = _run(adapter.fetch_snapshot())
    rows = asyncio.run(_collect_stream(adapter))
    row = rows[0]
    assert row.provenance.source_id == "test-src"
    assert row.provenance.source_row_ref == "42"
    assert row.provenance.source_snapshot_id == snapshot.snapshot_id
    assert len(row.provenance.source_row_hash) == 64


def test_stream_rows_provenance_bound_to_adapter_snapshot():
    """Row provenance must reference the adapter-generated snapshot_id, not an injected one."""
    xlsx = _make_xlsx([("A", 1, "100"), ("B", 2, "200")])
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {"etag": '"e"', "last_modified": ""}
    snapshot = _run(adapter.fetch_snapshot())
    rows = asyncio.run(_collect_stream(adapter))
    for row in rows:
        assert row.provenance.source_snapshot_id == snapshot.snapshot_id


def test_stream_rows_hash_stable():
    xlsx = _make_xlsx([("A", 42, "100")])
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {"etag": '"e"', "last_modified": ""}
    _run(adapter.fetch_snapshot())
    r1 = asyncio.run(_collect_stream(adapter))
    # Calling stream_rows again on the same snapshot yields same hashes
    r2 = asyncio.run(_collect_stream(adapter))
    assert r1[0].row_hash == r2[0].row_hash


def test_stream_rows_raises_before_connect():
    adapter = _adapter()
    with pytest.raises(RuntimeError, match="connect()"):
        asyncio.run(_collect_stream(adapter))


def test_stream_rows_raises_before_fetch_snapshot():
    """stream_rows() must fail if fetch_snapshot() was never called."""
    xlsx = _make_xlsx([("A", 1, "100")])
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {"etag": '"e"', "last_modified": ""}
    # connect() done (simulated), but fetch_snapshot() not called
    with pytest.raises(RuntimeError, match="fetch_snapshot()"):
        asyncio.run(_collect_stream(adapter))


def test_stream_rows_raises_on_duplicate_ids():
    """A source with duplicate IDs must raise before streaming begins."""
    xlsx = _make_xlsx([("A", 7, "100"), ("B", 7, "200")])
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {}
    with pytest.raises(ValueError, match="validation error"):
        _run(adapter.fetch_snapshot())


# ── HIGH 4: no silent row cap ─────────────────────────────────────────────────

def test_duplicate_detected_at_high_row_number():
    """
    A duplicate product ID at a high row number must NOT be silently skipped.
    Previously a _MAX_ROW cap silently ignored rows above position ~1000.
    """
    xlsx = _make_xlsx_with_late_duplicate(early_pid=42, late_row=1050)
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {}
    result = _run(adapter.validate_source())
    assert result.is_valid is False
    assert any("42" in e for e in result.errors)


def test_all_rows_returned_when_no_duplicates():
    """With 1000+ valid rows (no duplicates), all should be returned."""
    rows = [(f"Prod {i}", i, f"{i * 100}") for i in range(1, 1001)]
    xlsx = _make_xlsx(rows)
    adapter = _adapter()
    adapter._xlsx_bytes = xlsx
    adapter._meta = {"etag": '"e"', "last_modified": ""}
    _run(adapter.fetch_snapshot())
    streamed = asyncio.run(_collect_stream(adapter))
    assert len(streamed) == 1000


# ── MEDIUM 1: schema_hash covers all worksheets ───────────────────────────────

def test_schema_hash_differs_when_second_sheet_header_changes():
    """
    Changing a column header on a non-first worksheet must change schema_hash.
    Previously only the first sheet header was hashed.
    """
    # Two sheets with identical headers
    sheets_v1 = [
        [("A", 1, "100")],
        [("B", 2, "200")],
    ]
    # Second sheet gets a different header (we'll manipulate via openpyxl directly)
    import io
    from openpyxl import Workbook

    def _build(sheet2_header: str) -> bytes:
        wb = Workbook()
        ws1 = wb.active
        ws1.cell(row=1, column=1).value = "Label"
        ws1.cell(row=1, column=2).value = "Product ID"
        ws1.cell(row=1, column=3).value = "Price"
        ws1.cell(row=3, column=1).value = "A"
        ws1.cell(row=3, column=2).value = 1
        ws1.cell(row=3, column=3).value = "100"

        ws2 = wb.create_sheet()
        ws2.cell(row=1, column=1).value = "Label"
        ws2.cell(row=1, column=2).value = sheet2_header  # vary this
        ws2.cell(row=1, column=3).value = "Price"
        ws2.cell(row=3, column=1).value = "B"
        ws2.cell(row=3, column=2).value = 2
        ws2.cell(row=3, column=3).value = "200"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    xlsx_v1 = _build("Product ID")
    xlsx_v2 = _build("CHANGED_HEADER")

    adapter = _adapter()
    adapter._xlsx_bytes = xlsx_v1
    adapter._meta = {"etag": '"e1"', "last_modified": ""}
    snap1 = _run(adapter.fetch_snapshot())

    adapter2 = _adapter()
    adapter2._xlsx_bytes = xlsx_v2
    adapter2._meta = {"etag": '"e2"', "last_modified": ""}
    snap2 = _run(adapter2.fetch_snapshot())

    assert snap1.schema_hash != snap2.schema_hash


def test_schema_hash_stable_across_same_content():
    """Same XLSX content must produce the same schema_hash."""
    xlsx = _make_xlsx([("A", 1, "100")])
    adapter1 = _adapter()
    adapter1._xlsx_bytes = xlsx
    adapter1._meta = {"etag": '"e"', "last_modified": ""}
    snap1 = _run(adapter1.fetch_snapshot())

    adapter2 = _adapter()
    adapter2._xlsx_bytes = xlsx
    adapter2._meta = {"etag": '"e"', "last_modified": ""}
    snap2 = _run(adapter2.fetch_snapshot())

    assert snap1.schema_hash == snap2.schema_hash


# ── get_capabilities() ────────────────────────────────────────────────────────

def test_capabilities():
    adapter = _adapter()
    caps = adapter.get_capabilities()
    assert caps.supports_streaming is False   # XLSX reads full workbook into memory
    assert caps.supports_snapshots is True
    assert caps.supports_checkpointing is True
    assert caps.supports_incremental_sync is False
    assert caps.supports_deletions is False


# ── get_checkpoint() ──────────────────────────────────────────────────────────

def test_get_checkpoint_returns_etag():
    adapter = _adapter()
    adapter._xlsx_bytes = b"x"
    adapter._meta = {"etag": '"my-etag"', "last_modified": ""}
    cp = _run(adapter.get_checkpoint())
    assert cp is not None
    assert cp.checkpoint_value == '"my-etag"'
    assert cp.checkpoint_type == "etag"
    assert cp.source_id == "test-src"


def test_get_checkpoint_returns_none_when_no_etag():
    adapter = _adapter()
    adapter._xlsx_bytes = b"x"
    adapter._meta = {"etag": "", "last_modified": ""}
    with patch.object(adapter, "_fetch_meta_only", new=AsyncMock(return_value={"etag": "", "last_modified": ""})):
        cp = _run(adapter.get_checkpoint())
    assert cp is None


# ── HIGH 1: advance_checkpoint() persists via CheckpointRepository ────────────

def test_advance_checkpoint_persists_to_db():
    """advance_checkpoint(cp, db=session) must durably persist via CheckpointRepository."""
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from sqlalchemy.pool import StaticPool

    from app.a2.database import A2Base
    from app.a2.repositories.checkpoint_repository import CheckpointRepository
    from app.a2.repositories.source_repository import SourceRepository
    import app.a2.models.canonical_product  # noqa: F401
    import app.a2.models.source  # noqa: F401
    import app.a2.models.snapshot  # noqa: F401
    import app.a2.models.provenance  # noqa: F401
    import app.a2.models.checkpoint  # noqa: F401

    eng = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    A2Base.metadata.create_all(eng)
    Session = sessionmaker(bind=eng)
    db = Session()

    src_repo = SourceRepository(db)
    src_repo.create(
        source_id="test-src",
        source_type="nextcloud_xlsx",
        display_name="Test Source",
    )
    db.commit()

    adapter = _adapter()
    cp = SourceCheckpoint(
        source_id="test-src",
        checkpoint_value='"etag-persist"',
        checkpointed_at=datetime.now(tz=timezone.utc),
        checkpoint_type="etag",
    )
    _run(adapter.advance_checkpoint(cp, db=db))

    cp_repo = CheckpointRepository(db)
    saved = cp_repo.get("test-src")
    assert saved is not None
    assert saved.checkpoint_value == '"etag-persist"'
    assert saved.checkpoint_type == "etag"

    db.close()
    eng.dispose()


def test_advance_checkpoint_without_db_does_not_raise():
    """advance_checkpoint(cp, db=None) must log a warning but not raise."""
    adapter = _adapter()
    cp = SourceCheckpoint(
        source_id="test-src",
        checkpoint_value='"etag-nodb"',
        checkpointed_at=datetime.now(tz=timezone.utc),
        checkpoint_type="etag",
    )
    _run(adapter.advance_checkpoint(cp, db=None))  # must not raise


def test_advance_checkpoint_rollback_on_error():
    """advance_checkpoint must not silently swallow DB errors."""
    from unittest.mock import MagicMock
    from sqlalchemy.exc import SQLAlchemyError

    adapter = _adapter()
    cp = SourceCheckpoint(
        source_id="test-src",
        checkpoint_value='"etag-err"',
        checkpointed_at=datetime.now(tz=timezone.utc),
        checkpoint_type="etag",
    )
    mock_db = MagicMock()
    mock_db.get.side_effect = SQLAlchemyError("db error")

    with pytest.raises(SQLAlchemyError):
        _run(adapter.advance_checkpoint(cp, db=mock_db))


# ── SOURCE_TYPE constant ──────────────────────────────────────────────────────

def test_source_type_constant():
    assert NextcloudSourceAdapter.SOURCE_TYPE == "nextcloud_xlsx"
